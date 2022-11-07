import firebase_admin

from firebase_admin import db

from openpyxl import Workbook,load_workbook




cred_obj = firebase_admin.credentials.Certificate('info.json')
default_app = firebase_admin.initialize_app(cred_obj, {
	'databaseURL':'https://vindsql-14117-default-rtdb.firebaseio.com'
	})


ref = db.reference("Thermodynamics")


wb = load_workbook("Ther.xlsx")


wb.active = wb["Sheet1"]

ws=wb.active

key=['C','D','E','F','G']

for i in range(0,len(key)):
    for  j in range(1,29):
        ref1=ref.child("Model").child(ws[key[i]+'1'].value.replace(" ","")).child("value").push()
        ref1.child("Name").set(str(ws['A'+str(j)].value))
        ref1.child("Value").set(str(ws[str(key[i]+str(j))].value))
      
        
